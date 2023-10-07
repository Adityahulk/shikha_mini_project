import random
from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
import config
from bs4 import BeautifulSoup as bs
import requests
from selenium.webdriver.support.ui import WebDriverWait
from time import sleep
import openpyxl

headers = ["Perfume Name", "Perfume Url", "Image Url", "Brand Name", "Perfume Rating", "Top Notes", "Middle Notes", "Base Notes", "Longevity", "Sillage", "Gender", "Price Value"]

def use_proxy(perfume_url):
    # def getProxy():
    #     proxy = random.choice(config.proxy_list)
    #     return proxy

    t = len(perfume_url)
    print("hello")
    cnt = 1
    i=1
    start=i
    while i<1000:
        # proxy = getProxy()
        # http_proxy = "http://" + proxy
        # https_proxy = "https://" + proxy
        # print(http_proxy)
        # print(https_proxy)


        driver = None  # Initialize driver as None

        # Configure Selenium with the proxy
        # chrome_options = ChromeOptions()
        # chrome_options.add_argument(f'--proxy-server={https_proxy}')

        try:
            # Initialize the Chrome WebDriver with the configured proxy
            # driver = webdriver.Chrome(options=chrome_options)

            # Use Selenium to automate web interactions
            driver = webdriver.Chrome()
            driver.get("https://www.fragrantica.com" + perfume_url[i])
            print(perfume_url[i])
            sleep(3)
            page = driver.page_source
            # print(page)
            file_name = (str)(cnt) + ".html"
            with open(file_name,"w", encoding="utf-8") as f :
                f.write(page)

            page = open(file_name,'r',encoding="utf-8")
            soup = bs(page,'html.parser')
            # print(soup.prettify())
            pfname = soup.find('div',itemprop ='description')
            first_bold_element = pfname.find('b')
            perfume_name = first_bold_element.get_text()

            div_elements = soup.find_all('div', style = "display: flex; justify-content: center; text-align: center; flex-flow: wrap; align-items: flex-end; padding: 0.5rem;")
            dict = {}
            notes = ["Top Notes", "Middle Notes", "Base Notes"]
            ct = 0
            for div in div_elements:
                temp = []
                for it in div:
                    text = it.get_text()
                    temp.append(text)
                dict[notes[ct]] = temp
                ct += 1

            top_notes = dict["Top Notes"]
            middle_notes = dict["Middle Notes"]
            base_notes = dict["Base Notes"]

            img = soup.find('img', itemprop = "image")
            image = img['src']

            rating = soup.find('span', itemprop = 'ratingValue').get_text(strip=True)

            brand_div = soup.find('a', itemprop = "url")
            brand_name = brand_div.get_text(strip = True)

            mp = ["Longevity", "Sillage", "Gender", "Price Value"]
            mp_divs = soup.find_all('div', class_ = "cell small-12 medium-6")
            vote_dict = {}
            ct = 0
            for divs in mp_divs:
                div = divs.find_all('div',class_="grid-x grid-margin-x")
                if(len(div)!=0): 
                    val = []
                    for it in div:
                        tag = it.find_all('span')
                        temp = []
                        for t in tag:
                            temp.append(t.get_text(strip=True))
                        val.append(temp)
                    vote_dict[mp[ct]] = val
                    ct += 1


            Longevity = vote_dict["Longevity"]
            Sillage = vote_dict["Sillage"]
            Gender = vote_dict["Gender"]
            Price_value = vote_dict["Price Value"]
            # print(Longevity)
            workbook = openpyxl.load_workbook('data.xlsx')
            sheet = workbook["Sheet1"]
            data_row = [perfume_name, perfume_url[i], image, brand_name, rating, ", ".join(top_notes), ", ".join(middle_notes), ", ".join(base_notes), ", ".join(f'{item[0]}: {item[1]}' for item in Longevity), ", ".join(f'{item[0]}: {item[1]}' for item in Sillage), ", ".join(f'{item[0]}: {item[1]}' for item in Gender), ", ".join(f'{item[0]}: {item[1]}' for item in Price_value)]
            # # print(data_to_add)
            # # print("*****\n")
            sheet.append(data_row)
            workbook.save('data.xlsx')
            print("sucesss for" + perfume_url[i])
            i = i+1
            cnt = cnt + 1
        except Exception as e:
            print("There was an error","index",i,"\n")
            print(str(e))
            # config.proxy_list.pop(proxy)
            if len(config.proxy_list)==0:
                print("got the data till {start}")
                break
        finally:
            if driver is not None:
                driver.quit()
        # driver.close()
# Check if driver is not None before quitting
    print("****************************************************\n")
    print(i)
    print("****************************************************\n")

perfume_url = []
workbook = openpyxl.load_workbook("perfumes_names.xlsx")
worksheet = workbook['Sheet1']
for row in worksheet.iter_rows(values_only=True):
    for cell_value in row:
        perfume_url.append(cell_value)

workbook.close()

# Call the function with the list of URLs
use_proxy(perfume_url)